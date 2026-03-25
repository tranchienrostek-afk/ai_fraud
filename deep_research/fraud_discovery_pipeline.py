"""
AZINSU - AI Fraud Discovery Pipeline
Automated Graph Audit & Executive Report Generator
Chay: python fraud_discovery_pipeline.py
"""

import pandas as pd
from neo4j import GraphDatabase
from datetime import datetime
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── CAU HINH ─────────────────────────────────────────────────────
NEO4J_URI = "neo4j://127.0.0.1:7687"
NEO4J_USER = "neo4j"
NEO4J_PASS = "Chien@2022"
NEO4J_DB = "neo4j"

TODAY = datetime.now().strftime("%Y%m%d")
OUTPUT_FILE = f"Fraud_Discovery_Report_{TODAY}.xlsx"

# ── SCHEMA THUC TE ───────────────────────────────────────────────
# Person: user_id, identity_number(float), full_name, phone_number(float), email, salary, contract_level
# Claim:  claim_id, claim_date(string), amount, visit_type(float), diagnosis
# Hospital: hospital_code (ONLY - no name)
# Doctor: name
# Bank: account_number, beneficiary_name
# Insurer: insurer_id (no name)
# Expense: category (no item_name, no amount)
# Relationships: SUBMITTED, TREATED_AT, EXAMINED_BY, HANDLED_BY, PAID_TO, HAS_EXPENSE

COLOR_CRITICAL = "FF4444"
COLOR_HIGH = "FF8C00"
COLOR_MEDIUM = "FFD700"
COLOR_LOW = "90EE90"
COLOR_HEADER = "1B3A5C"
COLOR_SUBHEADER = "2E75B6"

# ── TRUY VAN CYPHER ──────────────────────────────────────────────

QUERIES = {
    "1_Petty_Fraud_Storm": {
        "title": "Truc loi vun (Petty Fraud Storm)",
        "description": "Ho so < 200k lap > 7 lan - lach nguong VAT, loi dung duyet tu dong",
        "severity": "CAO",
        "cypher": """
            MATCH (p:Person)-[:SUBMITTED]->(c:Claim)
            WHERE c.amount < 200000
            WITH p, count(c) AS so_ho_so,
                 sum(c.amount) AS tong_tien,
                 min(c.claim_date) AS tu_ngay,
                 max(c.claim_date) AS den_ngay,
                 collect(c.claim_date) AS dates
            WHERE so_ho_so > 7
            RETURN p.user_id AS user_id,
                   p.full_name AS ho_ten,
                   toString(toInteger(p.phone_number)) AS sdt,
                   so_ho_so,
                   round(tong_tien) AS tong_tien,
                   tu_ngay, den_ngay
            ORDER BY so_ho_so DESC
            LIMIT 100
        """,
    },
    "2_Facility_Loyalty": {
        "title": "Trung thanh CSYT bat thuong (Facility Loyalty)",
        "description": "Benh vat kham > 2 lan cung 1 co so - dau hieu nuoi ho so hoac thong dong",
        "severity": "CAO",
        "cypher": """
            MATCH (p:Person)-[:SUBMITTED]->(c:Claim)-[:TREATED_AT]->(h:Hospital)
            WHERE c.diagnosis =~ '(?i).*(vi\u00eam h\u1ecdng|vi\u00eam m\u0169i|c\u1ea3m c\u00fam|r\u1ed1i lo\u1ea1n ti\u00eau|vi\u00eam d\u1ea1 d\u00e0y|vi\u00eam ph\u1ebf qu\u1ea3n|vi\u00eam k\u1ebft m\u1ea1c).*'
            WITH p, h, c.diagnosis AS chan_doan,
                 count(c) AS so_lan_kham,
                 sum(c.amount) AS tong_tien
            WHERE so_lan_kham > 2
            RETURN p.user_id AS user_id,
                   p.full_name AS ho_ten,
                   h.hospital_code AS ma_benh_vien,
                   chan_doan,
                   so_lan_kham,
                   round(tong_tien) AS tong_tien
            ORDER BY so_lan_kham DESC
            LIMIT 100
        """,
    },
    "3_Phone_PII_Hub": {
        "title": "Cum dung chung SĐT (Phone PII Hub)",
        "description": "Nhom > 5 nguoi dung chung SDT - dau hieu duong day co to chuc",
        "severity": "CUC CAO",
        "cypher": """
            MATCH (p:Person)
            WHERE p.phone_number > 0
            WITH p.phone_number AS phone_raw,
                 collect(DISTINCT p.full_name) AS members,
                 collect(DISTINCT p.user_id) AS user_ids
            WHERE size(members) > 5
            WITH phone_raw, members, user_ids
            UNWIND user_ids AS uid
            OPTIONAL MATCH (pp:Person {user_id: uid})-[:SUBMITTED]->(c:Claim)
            WITH phone_raw, members,
                 count(c) AS total_claims,
                 sum(c.amount) AS total_amount
            RETURN toString(toInteger(phone_raw)) AS sdt,
                   size(members) AS so_nguoi,
                   total_claims AS tong_ho_so,
                   round(total_amount) AS tong_tien,
                   members[0..5] AS mau_ten
            ORDER BY so_nguoi DESC
            LIMIT 50
        """,
    },
    "4_Bank_Syndicate": {
        "title": "Mang luoi STK Ngan hang (Bank Syndicate)",
        "description": "Nhieu nguoi nhan tien ve cung 1 STK - trung tam tai chinh cua duong day",
        "severity": "CUC CAO",
        "cypher": """
            MATCH (p:Person)-[:SUBMITTED]->(c:Claim)-[:PAID_TO]->(b:Bank)
            WITH b, collect(DISTINCT p.full_name) AS users,
                 collect(DISTINCT p.user_id) AS user_ids,
                 count(c) AS total_claims,
                 sum(c.amount) AS total_payout
            WHERE size(users) > 3
            RETURN b.account_number AS stk,
                   b.beneficiary_name AS chu_tk,
                   size(users) AS so_nguoi,
                   total_claims AS tong_ho_so,
                   round(total_payout) AS tong_tien_nhan,
                   users[0..5] AS mau_ten
            ORDER BY so_nguoi DESC
            LIMIT 50
        """,
    },
    "5_Doctor_Hospital_Hotspot": {
        "title": "Cum Bac si - Benh vien nghi van",
        "description": "Bac si ky nhieu ho so bat thuong tai cung 1 benh vien",
        "severity": "CAO",
        "cypher": """
            MATCH (c:Claim)-[:EXAMINED_BY]->(d:Doctor),
                  (c)-[:TREATED_AT]->(h:Hospital)
            WITH d, h, count(c) AS so_ho_so,
                 sum(c.amount) AS tong_tien,
                 collect(DISTINCT c.diagnosis) AS danh_sach_chan_doan
            WHERE so_ho_so > 20
            RETURN d.name AS bac_si,
                   h.hospital_code AS ma_bv,
                   so_ho_so,
                   round(tong_tien) AS tong_tien,
                   size(danh_sach_chan_doan) AS so_loai_chan_doan,
                   danh_sach_chan_doan[0..3] AS mau_chan_doan
            ORDER BY so_ho_so DESC
            LIMIT 50
        """,
    },
    "6_Ghost_Clinic": {
        "title": "CSYT nghi van (Ghost Clinic)",
        "description": "BV co > 70% hoa don nho (< 200k) - dau hieu CSYT cap chung tu gia",
        "severity": "CAO",
        "cypher": """
            MATCH (h:Hospital)<-[:TREATED_AT]-(c:Claim)
            WITH h, count(c) AS total,
                 sum(CASE WHEN c.amount < 200000 THEN 1 ELSE 0 END) AS petty_count,
                 sum(c.amount) AS total_amount
            WHERE total > 10
            RETURN h.hospital_code AS ma_bv,
                   total AS tong_ho_so,
                   petty_count AS ho_so_nho,
                   round(petty_count * 100.0 / total) AS ti_le_nho_pct,
                   round(total_amount) AS tong_tien
            ORDER BY ti_le_nho_pct DESC
            LIMIT 30
        """,
    },
    "7_High_Freq_Claimant": {
        "title": "Nguoi nop ho so nhieu bat thuong",
        "description": "Khach hang nop > 10 ho so - dau hieu truc loi co he thong",
        "severity": "CAO",
        "cypher": """
            MATCH (p:Person)-[:SUBMITTED]->(c:Claim)
            WITH p, count(c) AS so_ho_so,
                 sum(c.amount) AS tong_tien,
                 collect(DISTINCT c.diagnosis) AS chan_doan_list,
                 min(c.claim_date) AS tu_ngay,
                 max(c.claim_date) AS den_ngay
            WHERE so_ho_so > 10
            RETURN p.user_id AS user_id,
                   p.full_name AS ho_ten,
                   toString(toInteger(p.phone_number)) AS sdt,
                   so_ho_so,
                   round(tong_tien) AS tong_tien,
                   size(chan_doan_list) AS so_loai_benh,
                   tu_ngay, den_ngay,
                   chan_doan_list[0..3] AS mau_chan_doan
            ORDER BY so_ho_so DESC
            LIMIT 50
        """,
    },
    "8_Multi_Hospital_Same_Diag": {
        "title": "Kham nhieu BV cung chan doan (Doctor Shopping)",
        "description": "Cung 1 benh, kham >= 3 BV khac nhau - dau hieu gom ho so",
        "severity": "TRUNG BINH",
        "cypher": """
            MATCH (p:Person)-[:SUBMITTED]->(c:Claim)-[:TREATED_AT]->(h:Hospital)
            WHERE c.diagnosis IS NOT NULL
            WITH p, c.diagnosis AS chan_doan,
                 collect(DISTINCT h.hospital_code) AS ds_bv,
                 count(c) AS so_ho_so,
                 sum(c.amount) AS tong_tien
            WHERE size(ds_bv) >= 3
            RETURN p.user_id AS user_id,
                   p.full_name AS ho_ten,
                   chan_doan,
                   size(ds_bv) AS so_bv_khac_nhau,
                   so_ho_so,
                   round(tong_tien) AS tong_tien,
                   ds_bv[0..5] AS mau_bv
            ORDER BY so_bv_khac_nhau DESC
            LIMIT 50
        """,
    },
}

# Summary queries
SUMMARY_QUERIES = {
    "overview": """
        MATCH (c:Claim)
        RETURN count(c) AS tong_claims,
               round(sum(c.amount)) AS tong_tien,
               round(avg(c.amount)) AS amount_tb
    """,
    "top_claimants": """
        MATCH (p:Person)-[:SUBMITTED]->(c:Claim)
        WITH p, count(c) AS num_claims, sum(c.amount) AS total_amount
        WHERE num_claims > 5
        RETURN p.user_id AS user_id,
               p.full_name AS ho_ten,
               toString(toInteger(p.phone_number)) AS sdt,
               num_claims AS so_ho_so,
               round(total_amount) AS tong_tien
        ORDER BY total_amount DESC
        LIMIT 30
    """,
    "agent_risk": """
        MATCH (c:Claim)-[:HANDLED_BY]->(ins:Insurer)
        WITH ins, count(c) AS total,
             sum(c.amount) AS total_amount,
             sum(CASE WHEN c.amount < 200000 THEN 1 ELSE 0 END) AS petty
        WHERE total > 5
        RETURN ins.insurer_id AS dai_ly_id,
               total AS tong_ho_so,
               petty AS ho_so_nho,
               round(petty * 100.0 / total) AS ti_le_nho_pct,
               round(total_amount) AS tong_tien
        ORDER BY tong_ho_so DESC
        LIMIT 20
    """,
}


# ── ENGINE ───────────────────────────────────────────────────────

def connect_neo4j():
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASS))
    driver.verify_connectivity()
    print("[OK] Ket noi Neo4j thanh cong")
    return driver


def run_query(driver, cypher, name=""):
    try:
        records, _, _ = driver.execute_query(cypher, database_=NEO4J_DB)
        if not records:
            print(f"  -> {name}: 0 dong")
            return pd.DataFrame()
        rows = []
        for r in records:
            row = {}
            for k in r.keys():
                v = r[k]
                # Convert neo4j lists to strings for Excel
                if isinstance(v, list):
                    v = " | ".join(str(x) for x in v)
                row[k] = v
            rows.append(row)
        df = pd.DataFrame(rows)
        print(f"  -> {name}: {len(df)} dong")
        return df
    except Exception as e:
        print(f"  [LOI] {name}: {e}")
        return pd.DataFrame()


def severity_color(severity):
    if "CUC" in severity:
        return COLOR_CRITICAL
    if "CAO" in severity:
        return COLOR_HIGH
    if "TRUNG" in severity:
        return COLOR_MEDIUM
    return COLOR_LOW


def write_excel_report(audit_results, summary_data):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        # ── SHEET: TONG QUAN ──────────────────────────────────
        ws = writer.book.create_sheet("TONG_QUAN", 0)
        ws.sheet_properties.tabColor = COLOR_HEADER

        # Title row
        ws.merge_cells("A1:G1")
        c = ws["A1"]
        c.value = "BAO CAO PHAT HIEN GIAN LAN BAO HIEM - AZINSU"
        c.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 45

        # Subtitle
        ws.merge_cells("A2:G2")
        c = ws["A2"]
        c.value = f"Bao cao tu dong | Ngay: {datetime.now().strftime('%d/%m/%Y %H:%M')} | AI Fraud Discovery Pipeline v2"
        c.font = Font(name="Arial", size=10, italic=True, color="666666")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 25

        # Overview stats
        ov = summary_data.get("overview")
        if ov is not None and not ov.empty:
            row = ov.iloc[0]
            stats = [
                ("Tong so ho so (Claims)", row.get("tong_claims", 0)),
                ("Tong gia tri boi thuong (VND)", f"{row.get('tong_tien', 0):,.0f}"),
                ("So tien TB/ho so (VND)", f"{row.get('amount_tb', 0):,.0f}"),
            ]
            for i, (label, value) in enumerate(stats):
                r = 4 + i
                ws[f"A{r}"] = label
                ws[f"A{r}"].font = Font(name="Arial", size=11, bold=True)
                ws[f"C{r}"] = str(value)
                ws[f"C{r}"].font = Font(name="Arial", size=11, color=COLOR_CRITICAL)
                ws[f"C{r}"].alignment = Alignment(horizontal="right")

        # Audit summary table
        r = 9
        headers = ["STT", "Quy tac Audit", "Muc do", "So phat hien", "Mo ta"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=r, column=i + 1, value=h)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for idx, (key, cfg) in enumerate(QUERIES.items()):
            r += 1
            df = audit_results.get(key, pd.DataFrame())
            count = len(df)
            sev = cfg["severity"]

            ws.cell(row=r, column=1, value=idx + 1).border = thin
            c = ws.cell(row=r, column=2, value=cfg["title"])
            c.border = thin
            c.font = Font(name="Arial", size=10, bold=True)

            sc = ws.cell(row=r, column=3, value=sev)
            sc.fill = PatternFill("solid", fgColor=severity_color(sev))
            sc.font = Font(name="Arial", size=10, bold=True)
            sc.border = thin
            sc.alignment = Alignment(horizontal="center")

            cc = ws.cell(row=r, column=4, value=count)
            cc.font = Font(name="Arial", size=12, bold=True,
                           color=COLOR_CRITICAL if count > 0 else "000000")
            cc.border = thin
            cc.alignment = Alignment(horizontal="center")

            dc = ws.cell(row=r, column=5, value=cfg["description"])
            dc.border = thin
            dc.alignment = Alignment(wrap_text=True)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 65

        # ── DETAIL SHEETS ────────────────────────────────────
        for key, cfg in QUERIES.items():
            df = audit_results.get(key, pd.DataFrame())
            sheet_name = key[:31]

            if df.empty:
                df = pd.DataFrame({"Thong_bao": ["Khong co ket qua"]})

            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            ws = writer.sheets[sheet_name]
            ws.sheet_properties.tabColor = severity_color(cfg["severity"])

            max_col = max(len(df.columns), 1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            hc = ws.cell(row=1, column=1)
            hc.value = f"{cfg['title']}  |  Muc do: {cfg['severity']}  |  {cfg['description']}"
            hc.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
            hc.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[1].height = 35

            # Style column headers (row 3)
            for col in range(1, max_col + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Auto-width columns
            for col in range(1, max_col + 1):
                cl = get_column_letter(col)
                max_len = len(str(ws.cell(row=3, column=col).value or ""))
                for row_idx in range(4, min(len(df) + 4, 25)):
                    val_len = len(str(ws.cell(row=row_idx, column=col).value or ""))
                    if val_len > max_len:
                        max_len = val_len
                ws.column_dimensions[cl].width = min(max_len + 4, 50)

            # Data cell borders
            for row_idx in range(4, len(df) + 4):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin
                    cell.alignment = Alignment(wrap_text=True)

        # ── TOP CLAIMANTS SHEET ──────────────────────────────
        for sname, skey, color, title in [
            ("TOP_BOI_THUONG", "top_claimants", COLOR_CRITICAL, "TOP KHACH HANG BOI THUONG CAO NHAT"),
            ("DAI_LY", "agent_risk", COLOR_HIGH, "PHAN TICH THEO DAI LY BAO HIEM"),
        ]:
            sdf = summary_data.get(skey, pd.DataFrame())
            if sdf.empty:
                continue
            sdf.to_excel(writer, sheet_name=sname, index=False, startrow=2)
            ws = writer.sheets[sname]
            ws.sheet_properties.tabColor = color
            mc = len(sdf.columns)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
            hc = ws.cell(row=1, column=1)
            hc.value = title
            hc.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor=color)
            for col in range(1, mc + 1):
                c = ws.cell(row=3, column=col)
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                c.border = thin

    print(f"\n[OK] Xuat bao cao: {OUTPUT_FILE}")


# ── MAIN ─────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  AZINSU - AI FRAUD DISCOVERY PIPELINE v2")
    print("=" * 60)

    # Connect
    print("\n[1/3] Ket noi Neo4j...")
    try:
        driver = connect_neo4j()
    except Exception as e:
        print(f"[LOI] Khong the ket noi: {e}")
        sys.exit(1)

    # Run audit rules
    print("\n[2/3] Chay 8 Quy tac Audit...")
    audit_results = {}
    for key, cfg in QUERIES.items():
        print(f"\n  [{cfg['severity']}] {cfg['title']}")
        audit_results[key] = run_query(driver, cfg["cypher"], name=key)

    # Run summary queries
    print("\n  --- Truy van tong hop ---")
    summary_data = {}
    for key, cypher in SUMMARY_QUERIES.items():
        summary_data[key] = run_query(driver, cypher, name=key)

    # Export
    print("\n[3/3] Xuat bao cao Excel...")
    write_excel_report(audit_results, summary_data)

    # Summary
    print("\n" + "=" * 60)
    print("  KET QUA TONG HOP")
    print("=" * 60)
    total = 0
    for key, cfg in QUERIES.items():
        df = audit_results.get(key, pd.DataFrame())
        n = len(df)
        total += n
        flag = " ***" if n > 0 else ""
        print(f"  {cfg['title']}: {n} phat hien{flag}")

    print(f"\n  TONG CONG: {total} phat hien")
    print(f"  FILE: {OUTPUT_FILE}")
    print("=" * 60)

    driver.close()


if __name__ == "__main__":
    main()
